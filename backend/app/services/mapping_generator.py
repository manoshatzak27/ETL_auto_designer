"""
Generates variable_mapping.csv, value_mapping.csv, variable_value_mapping.csv
and custom_mappings.csv from the user's concept decisions.

Decision structure per variable:
{
  "strategy": "map_variable" | "map_values" | "map_both" | "skip",
  "variable_concept": {"concept_id": int, "concept_name": str,
                       # optional fields, present when concept was created via
                       # the "Create custom concept" form:
                       "concept_code": str, "domain_id_str": str,
                       "vocabulary_id": str, "concept_class_id": str,
                       "is_custom": True} | null,
  "value_concepts": {
      "<source_value>": {... same shape ...}
  }
}

Strategies:
  map_variable  → variable_mapping.csv only (numeric variables)
  map_values    → variable_value_mapping.csv (categorical: each variable+value IS its own concept)
  map_both      → variable_mapping.csv + value_mapping.csv
                  (categorical: variable has a concept + each value has a value_as_concept_id)
  skip          → omitted from all files
"""
import io
import os
from datetime import date
from pathlib import Path
from typing import Callable
import pandas as pd
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

VALID_START = "1970-01-01"
VALID_END = "2099-12-31"
CUSTOM_CONCEPT_THRESHOLD = 2_000_000_000
DEFAULT_CUSTOM_VOCABULARY = "CUSTOM"


def _base_row(variable: str, concept: dict, domain_id: int | None = None) -> dict:
    return {
        "variable_source_code": variable,
        "source_code_description": concept.get("description", ""),
        "target_concept_id": concept["concept_id"],
        "target_concept_name": concept.get("concept_name", ""),
        "domain_id": domain_id if domain_id is not None else "",
        "valid_start_date": VALID_START,
        "valid_end_date": VALID_END,
        "invalid_reason": "",
    }


def _value_row(variable: str, value: str, concept: dict, domain_id: int | None = None) -> dict:
    return {
        "variable_source_code": variable,
        "value_source_code": value,
        "source_code_description": concept.get("description", ""),
        "target_concept_id": concept["concept_id"],
        "target_concept_name": concept.get("concept_name", ""),
        "domain_id": domain_id if domain_id is not None else "",
        "valid_start_date": VALID_START,
        "valid_end_date": VALID_END,
        "invalid_reason": "",
    }


def generate_mapping_csvs(
    concept_decisions: dict,
    output_dir: str,
    custom_vocabulary_id: str = DEFAULT_CUSTOM_VOCABULARY,
) -> dict[str, str]:
    """
    Generates the 4 mapping CSVs from concept_decisions dict.
    `custom_vocabulary_id` is the project-level vocab id that custom concepts
    (id >= 2_000_000_000) are tagged with when the concept dict doesn't carry
    one of its own.
    Returns a dict mapping key → absolute file path for each CSV produced.
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    variable_rows: list[dict] = []
    value_rows: list[dict] = []
    var_value_rows: list[dict] = []
    custom_rows: list[dict] = []
    unit_rows: list[dict] = []
    route_rows: list[dict] = []
    modifier_rows: list[dict] = []
    condition_status_rows: list[dict] = []
    qualifier_rows: list[dict] = []
    operator_rows: list[dict] = []

    for variable, decision in concept_decisions.items():
        if decision.get("enabled") is False:
            continue

        strategy = decision.get("strategy", "skip")
        if strategy == "skip":
            continue

        var_concept = decision.get("variable_concept") or {}
        val_concepts: dict = decision.get("value_concepts") or {}
        domain_id: int | None = decision.get("domain_id") or None

        # --- variable_mapping.csv ---
        if strategy in ("map_variable", "map_both") and var_concept.get("concept_id"):
            variable_rows.append(_base_row(variable, var_concept, domain_id=domain_id))
            if var_concept["concept_id"] >= CUSTOM_CONCEPT_THRESHOLD:
                custom_rows.append(_custom_row(var_concept, custom_vocabulary_id))

        # --- variable_value_mapping.csv  (categorical: variable+value = concept) ---
        if strategy == "map_values":
            for val, vc in val_concepts.items():
                if vc.get("concept_id"):
                    var_value_rows.append(_value_row(variable, val, vc, domain_id=vc.get("domain_id")))
                    if vc["concept_id"] >= CUSTOM_CONCEPT_THRESHOLD:
                        custom_rows.append(_custom_row(vc, custom_vocabulary_id))

        # --- value_mapping.csv  (value_as_concept_id alongside variable concept) ---
        if strategy == "map_both":
            for val, vc in val_concepts.items():
                if vc.get("concept_id"):
                    value_rows.append(_value_row(variable, val, vc, domain_id=domain_id))
                    if vc["concept_id"] >= CUSTOM_CONCEPT_THRESHOLD:
                        custom_rows.append(_custom_row(vc, custom_vocabulary_id))

        # --- unit_mapping.csv  (measurement/observation/drug dose unit_source_value → unit_concept_id) ---
        unit_mapping = decision.get("unit_mapping") or {}
        unit_col: str | None = unit_mapping.get("unit_col")
        unit_concepts: dict = unit_mapping.get("unit_concepts") or {}
        if unit_col and unit_concepts:
            for unit_val, unit_cid in unit_concepts.items():
                if unit_cid:
                    unit_rows.append({
                        "variable_source_code": variable,
                        "unit_col": unit_col,
                        "unit_source_value": unit_val,
                        "unit_concept_id": int(unit_cid),
                    })

        # --- route_mapping.csv  (drug_exposure route_source_value → route_concept_id) ---
        route_mapping = decision.get("route_mapping") or {}
        route_col: str | None = route_mapping.get("route_col")
        route_concepts: dict = route_mapping.get("route_concepts") or {}
        if route_col and route_concepts:
            for route_val, route_cid in route_concepts.items():
                if route_cid:
                    route_rows.append({
                        "variable_source_code": variable,
                        "route_col": route_col,
                        "route_source_value": route_val,
                        "route_concept_id": int(route_cid),
                    })

        # --- modifier_mapping.csv  (procedure_occurrence modifier_source_value → modifier_concept_id) ---
        modifier_mapping = decision.get("modifier_mapping") or {}
        modifier_col: str | None = modifier_mapping.get("modifier_col")
        modifier_concepts: dict = modifier_mapping.get("modifier_concepts") or {}
        if modifier_col and modifier_concepts:
            for modifier_val, modifier_cid in modifier_concepts.items():
                if modifier_cid:
                    modifier_rows.append({
                        "variable_source_code": variable,
                        "modifier_col": modifier_col,
                        "modifier_source_value": modifier_val,
                        "modifier_concept_id": int(modifier_cid),
                    })

        # --- condition_status_mapping.csv  (condition_occurrence condition_status_source_value → condition_status_concept_id) ---
        condition_status_mapping = decision.get("condition_status_mapping") or {}
        condition_status_col: str | None = condition_status_mapping.get("condition_status_col")
        condition_status_concepts: dict = condition_status_mapping.get("condition_status_concepts") or {}
        if condition_status_col and condition_status_concepts:
            for status_val, status_cid in condition_status_concepts.items():
                if status_cid:
                    condition_status_rows.append({
                        "variable_source_code": variable,
                        "condition_status_col": condition_status_col,
                        "condition_status_source_value": status_val,
                        "condition_status_concept_id": int(status_cid),
                    })

        # --- qualifier_mapping.csv  (observation qualifier_source_value → qualifier_concept_id) ---
        qualifier_mapping = decision.get("qualifier_mapping") or {}
        qualifier_col: str | None = qualifier_mapping.get("qualifier_col")
        qualifier_concepts: dict = qualifier_mapping.get("qualifier_concepts") or {}
        if qualifier_col and qualifier_concepts:
            for qualifier_val, qualifier_cid in qualifier_concepts.items():
                if qualifier_cid:
                    qualifier_rows.append({
                        "variable_source_code": variable,
                        "qualifier_col": qualifier_col,
                        "qualifier_source_value": qualifier_val,
                        "qualifier_concept_id": int(qualifier_cid),
                    })

        # --- operator_mapping.csv  (measurement operator_source_value → operator_concept_id) ---
        operator_mapping = decision.get("operator_mapping") or {}
        operator_col: str | None = operator_mapping.get("operator_col")
        operator_concepts: dict = operator_mapping.get("operator_concepts") or {}
        if operator_col and operator_concepts:
            for operator_val, operator_cid in operator_concepts.items():
                if operator_cid:
                    operator_rows.append({
                        "variable_source_code": variable,
                        "operator_col": operator_col,
                        "operator_source_value": operator_val,
                        "operator_concept_id": int(operator_cid),
                    })

    files: dict[str, str] = {}

    if variable_rows:
        df = pd.DataFrame(variable_rows).sort_values("variable_source_code")
        p = str(output_path / "variable_mapping.csv")
        df.to_csv(p, index=False, encoding="utf-8")
        files["variable_mapping"] = p

    if value_rows:
        df = pd.DataFrame(value_rows).sort_values(["variable_source_code", "value_source_code"])
        p = str(output_path / "value_mapping.csv")
        df.to_csv(p, index=False, encoding="utf-8")
        files["value_mapping"] = p

    if var_value_rows:
        df = pd.DataFrame(var_value_rows).sort_values(["variable_source_code", "value_source_code"])
        p = str(output_path / "variable_value_mapping.csv")
        df.to_csv(p, index=False, encoding="utf-8")
        files["variable_value_mapping"] = p

    if custom_rows:
        seen: set[int] = set()
        unique = [r for r in custom_rows if r["concept_id"] not in seen and not seen.add(r["concept_id"])]
        df = pd.DataFrame(unique).sort_values("concept_id")
        p = str(output_path / "custom_mappings.csv")
        df.to_csv(p, index=False, encoding="utf-8")
        files["custom_mappings"] = p

    if unit_rows:
        df = pd.DataFrame(unit_rows).sort_values(["variable_source_code", "unit_source_value"])
        p = str(output_path / "unit_mapping.csv")
        df.to_csv(p, index=False, encoding="utf-8")
        files["unit_mapping"] = p

    if route_rows:
        df = pd.DataFrame(route_rows).sort_values(["variable_source_code", "route_source_value"])
        p = str(output_path / "route_mapping.csv")
        df.to_csv(p, index=False, encoding="utf-8")
        files["route_mapping"] = p

    if modifier_rows:
        df = pd.DataFrame(modifier_rows).sort_values(["variable_source_code", "modifier_source_value"])
        p = str(output_path / "modifier_mapping.csv")
        df.to_csv(p, index=False, encoding="utf-8")
        files["modifier_mapping"] = p

    if condition_status_rows:
        df = pd.DataFrame(condition_status_rows).sort_values(["variable_source_code", "condition_status_source_value"])
        p = str(output_path / "condition_status_mapping.csv")
        df.to_csv(p, index=False, encoding="utf-8")
        files["condition_status_mapping"] = p

    if qualifier_rows:
        df = pd.DataFrame(qualifier_rows).sort_values(["variable_source_code", "qualifier_source_value"])
        p = str(output_path / "qualifier_mapping.csv")
        df.to_csv(p, index=False, encoding="utf-8")
        files["qualifier_mapping"] = p

    if operator_rows:
        df = pd.DataFrame(operator_rows).sort_values(["variable_source_code", "operator_source_value"])
        p = str(output_path / "operator_mapping.csv")
        df.to_csv(p, index=False, encoding="utf-8")
        files["operator_mapping"] = p

    return files


def _custom_row(concept: dict, custom_vocabulary_id: str = DEFAULT_CUSTOM_VOCABULARY) -> dict:
    # Custom concepts created via the form carry `domain_id_str` (e.g. "Observation"),
    # plus optional concept_code / concept_class_id. Anything missing falls back to
    # sensible defaults so a bare manual-ID entry still produces a valid row.
    return {
        "concept_id": concept["concept_id"],
        "concept_name": concept.get("concept_name", ""),
        "domain_id": concept.get("domain_id_str") or "Observation",
        "vocabulary_id": concept.get("vocabulary_id") or custom_vocabulary_id,
        "concept_class_id": concept.get("concept_class_id") or "Clinical Finding",
        "standard_concept": "S",
        "concept_code": concept.get("concept_code") or str(concept["concept_id"]),
        "valid_start_date": VALID_START,
        "valid_end_date": VALID_END,
        "invalid_reason": "",
    }


DOMAIN_LABELS = {
    1: "Measurement",
    2: "Observation",
    3: "Drug Exposure",
    4: "Procedure Occurrence",
    5: "Condition Occurrence",
}

STRATEGY_LABELS = {
    "skip": "Skip",
    "map_variable": "Map Variable",
    "map_values": "Map Values",
    "map_both": "Map Variable + Values",
}


def _fixed_concept(col: str | None, concepts: dict) -> tuple[str, int] | None:
    """Mirrors the frontend's getFixedConcept(): unit/route mappings are only
    a single variable-level concept when no source column is chosen and exactly
    one positive concept id is stored (keyed by its own concept name). Once a
    source column is chosen, `concepts` is instead keyed by that column's
    per-row source values, so there's no single id/name to show here."""
    if col:
        return None
    entries = [(k, v) for k, v in (concepts or {}).items() if isinstance(v, (int, float)) and v > 0]
    if len(entries) != 1:
        return None
    return entries[0]


def _autosize_columns(ws, df: pd.DataFrame) -> None:
    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].tolist()]) if len(df) else len(str(col))
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(60, max(10, max_len + 2))
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    ws.freeze_panes = "A2"


def generate_mapping_summary_excel(
    concept_decisions: dict,
    lookup_concept_name: Callable[[int], str] | None = None,
    column_values: dict[str, list[str]] | None = None,
) -> io.BytesIO:
    """
    Builds a human-readable Excel workbook summarizing every variable's mapping
    choices — unlike generate_mapping_csvs(), this includes every variable
    regardless of strategy/enabled state, for review rather than ETL loading.
    Variables are listed in `concept_decisions`' own iteration order — pass it
    in already ordered (e.g. by source-file column order) if that matters.

    `lookup_concept_name` resolves names for concept ids that are stored bare
    (per-column unit/route mappings only keep the id, keyed by source value);
    variable- and value-level concepts already carry their name inline.

    `column_values` (variable -> every distinct source value it has, mapped or
    not) makes map_values/map_both variables list every value, not just the
    ones the user assigned a concept to. Falls back to only the mapped values
    when a variable is missing from it (e.g. its source file is gone).

    Sheets:
      Variables & Values  one row per variable (every strategy/enabled state),
                           or one row per value for map_values/map_both
                           — variable-level columns repeat on every value row
      Unit & Route Values one row per distinct value of a column-mode unit/route
                           source column (mapped or not); variable name repeats
                           the same way as in Variables & Values
    """
    name_of = lookup_concept_name or (lambda _cid: "")
    column_values = column_values or {}

    rows: list[dict] = []
    unit_route_rows: list[dict] = []
    # Track where things land so the two sheets can be cross-linked afterwards:
    # each variable's own row in Variables & Values, and the row of its Unit/
    # Route header (if any) in Unit & Route Values. Excel rows are 1-indexed
    # with row 1 being the header, so a value's row is its 0-based list index + 2.
    var_row_in_main: dict[str, int] = {}
    unit_header_row: dict[str, int] = {}
    route_header_row: dict[str, int] = {}

    for variable, decision in concept_decisions.items():
        decision = decision or {}
        strategy = decision.get("strategy") or "skip"
        var_concept = decision.get("variable_concept") or {}
        domain_id = decision.get("domain_id")

        unit_mapping = decision.get("unit_mapping") or {}
        unit_col = unit_mapping.get("unit_col")
        unit_concepts = unit_mapping.get("unit_concepts") or {}
        unit_fixed = _fixed_concept(unit_col, unit_concepts)

        route_mapping = decision.get("route_mapping") or {}
        route_col = route_mapping.get("route_col")
        route_concepts = route_mapping.get("route_concepts") or {}
        route_fixed = _fixed_concept(route_col, route_concepts)

        # Column-mode unit/route entries for this variable — every distinct source
        # value of the chosen column (mapped or not), same "list everything"
        # approach as the value rows above. Each kind gets its own header row
        # (Kind set, no value), then its values start one row down; the variable
        # name is only stamped on the very first row of the whole group.
        entries: list[dict] = []
        for kind, col, mapping in (("Unit", unit_col, unit_concepts), ("Route", route_col, route_concepts)):
            if not col:
                continue
            header_row = len(unit_route_rows) + len(entries) + 2
            if kind == "Unit":
                unit_header_row[variable] = header_row
            else:
                route_header_row[variable] = header_row
            known = column_values.get(col)
            all_vals = sorted(set(known) | set(mapping.keys())) if known is not None else sorted(mapping.keys())
            entries.append({"Kind": kind, "Source Value": "", "Concept ID": "", "Concept Name": ""})
            for val in all_vals:
                cid = mapping.get(val)
                entries.append({
                    "Kind": "",
                    "Source Value": val,
                    "Concept ID": cid or "",
                    "Concept Name": name_of(cid) if cid else "",
                })
        for i, entry in enumerate(entries):
            unit_route_rows.append({"Variable": variable if i == 0 else "", **entry})

        base_row = {
            "Variable": variable,
            "Included": "No" if decision.get("enabled") is False else "Yes",
            "Strategy": STRATEGY_LABELS.get(strategy, strategy),
            "Domain": DOMAIN_LABELS.get(domain_id, ""),
            "Unit Concept ID": unit_fixed[1] if unit_fixed else "",
            "Unit Concept Name": unit_fixed[0] if unit_fixed else "",
            "Unit Source Column": unit_col or "",
            "Route Concept ID": route_fixed[1] if route_fixed else "",
            "Route Concept Name": route_fixed[0] if route_fixed else "",
            "Route Source Column": route_col or "",
            "Type Concept ID": decision.get("type_concept_id") or "",
            "Type Concept Name": decision.get("type_concept_name") or "",
            "Start Datetime Column": decision.get("start_datetime_col") or "",
            "End Datetime Column": decision.get("end_datetime_col") or "",
            "Datetime Format": decision.get("datetime_format") or "",
        }

        mapped_values: list[tuple[str, dict]] = []
        if strategy in ("map_values", "map_both"):
            value_concepts = decision.get("value_concepts") or {}
            known = column_values.get(variable)
            # List every distinct source value the column actually has (mapped or
            # not); if we couldn't read the source file, fall back to just the
            # values that got mapped so nothing is silently dropped.
            all_values = sorted(set(known) | set(value_concepts.keys())) if known is not None \
                else sorted(value_concepts.keys())
            mapped_values = [(val, value_concepts.get(val) or {}) for val in all_values]

        # Row 1 always carries the variable itself: its own concept (map_variable/
        # map_both) or a blank Concept ID/Name (map_values has no variable-level
        # concept, skip/unmapped have none set). Any mapped values follow one row
        # down each, with Source holding the source value and Variable/Included/
        # Strategy left blank so they aren't retyped per value.
        var_row_in_main[variable] = len(rows) + 2
        rows.append({
            **base_row,
            "Source": "",
            "Concept ID": var_concept.get("concept_id", ""),
            "Concept Name": var_concept.get("concept_name", ""),
        })
        for val, vc in mapped_values:
            rows.append({
                **base_row,
                # Every per-variable column (identity + unit/route/type/datetime)
                # is only typed once, on row 1 — value rows carry just the value.
                "Variable": "",
                "Included": "",
                "Strategy": "",
                "Unit Concept ID": "",
                "Unit Concept Name": "",
                "Unit Source Column": "",
                "Route Concept ID": "",
                "Route Concept Name": "",
                "Route Source Column": "",
                "Type Concept ID": "",
                "Type Concept Name": "",
                "Start Datetime Column": "",
                "End Datetime Column": "",
                "Datetime Format": "",
                # The value's own domain (falling back to the variable's) takes over
                # the shared Domain column here, same as Concept ID/Name above.
                "Domain": DOMAIN_LABELS.get(vc.get("domain_id", domain_id), ""),
                "Source": val,
                "Concept ID": vc.get("concept_id", ""),
                "Concept Name": vc.get("concept_name", ""),
            })

    main_cols = [
        "Variable", "Included", "Strategy", "Domain",
        "Source", "Concept ID", "Concept Name",
        "Unit Concept ID", "Unit Concept Name", "Unit Source Column",
        "Route Concept ID", "Route Concept Name", "Route Source Column",
        "Type Concept ID", "Type Concept Name",
        "Start Datetime Column", "End Datetime Column", "Datetime Format",
    ]
    unit_route_cols = ["Variable", "Kind", "Source Value", "Concept ID", "Concept Name"]

    df_main = pd.DataFrame(rows, columns=main_cols)
    # Not sorted — rows are already grouped per variable (in concept_decisions'
    # own order) with the variable name blanked after the first entry; sorting
    # by Variable would scatter those blanks and break the grouping.
    df_unit_route = pd.DataFrame(unit_route_rows, columns=unit_route_cols)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_main.to_excel(writer, sheet_name="Variables & Values", index=False)
        df_unit_route.to_excel(writer, sheet_name="Unit & Route Values", index=False)
        _autosize_columns(writer.sheets["Variables & Values"], df_main)
        _autosize_columns(writer.sheets["Unit & Route Values"], df_unit_route)

        # Link each variable's Unit/Route Source Column cell to where that
        # column-mode mapping's values actually live on the other sheet.
        ws_main = writer.sheets["Variables & Values"]
        unit_col_idx = main_cols.index("Unit Source Column") + 1
        route_col_idx = main_cols.index("Route Source Column") + 1
        hyperlink_font = Font(color="0563C1", underline="single")
        for variable, main_row in var_row_in_main.items():
            for header_rows, col_idx in ((unit_header_row, unit_col_idx), (route_header_row, route_col_idx)):
                target_row = header_rows.get(variable)
                if target_row is None:
                    continue
                cell = ws_main.cell(row=main_row, column=col_idx)
                # A plain string assigns to Hyperlink.target, which openpyxl always
                # serializes as an *external* relationship — Excel then can't resolve
                # a target starting with "#" and the link renders styled but dead.
                # Same-workbook jumps need `location` instead, with no target/relationship.
                cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'Unit & Route Values'!A{target_row}")
                cell.font = hyperlink_font
    buf.seek(0)
    return buf
